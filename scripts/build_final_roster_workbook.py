"""Build a private final-roster reconciliation workbook for QC review.

This is an offline transformation. It replaces the student and
student-assignment sheets with the Aug 25 enrollment roster, using the
existing Canvas/SIS reconciliation for section and assignment context. It
does not upload data or activate a Supabase batch.
"""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_SOURCE = Path(
    "outputs/ucsp_roster_reconciliation_20260824/"
    "FEU_HS_Teacher_Performance_Evaluation_Roster_UCSP_Reconciled_SectionsRefreshed.xlsx"
)
DEFAULT_ENROLLMENT = Path(
    Path.home() / "Downloads/"
    "FEU HS Enrolled as of Aug 25_SY26-27(3).xlsx"
)
DEFAULT_OUTPUT = Path("outputs/FEU_TPE_Final_Q1_2026-2027.xlsx")


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def grade_number(value: object) -> int:
    match = re.search(r"\d+", clean(value))
    if not match:
        raise SystemExit(f"Could not determine grade level from {value!r}")
    return int(match.group())


def canonical_section_code(section_name: object, existing_code: object = "") -> str:
    """Return the institutional code for a Canvas section long name.

    The prior reconciliation abbreviated General Studies as GE and assigned
    Engineering Science 1 the same code as Engineering Science 2. The SIS
    export establishes GS and the numbered ES codes as the authoritative form.
    """
    name = clean(section_name)
    match = re.fullmatch(r"11 General Studies (\d+)", name)
    if match:
        return f"11GS{int(match.group(1)):02d}"
    match = re.fullmatch(r"11 Engineering Science (\d+)", name)
    if match:
        return f"11ES{int(match.group(1)):02d}"
    return clean(existing_code)


def records(sheet) -> tuple[list[str], list[dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean(value) for value in rows[0]]
    return headers, [dict(zip(headers, row)) for row in rows[1:] if any(row)]


def replace_sheet(workbook, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    if name in workbook.sheetnames:
        del workbook[name]
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def fresh_qc(
    students: list[dict[str, object]],
    assignments: list[dict[str, object]],
    student_assignments: list[dict[str, object]],
    section_codes: set[str],
) -> tuple[list[str], list[dict[str, object]]]:
    """Create QC findings from the rebuilt rows only.

    The previous reconciliation's QC sheet is historical evidence, not a
    current validation result. Recomputing here prevents stale students and
    already-resolved substitution findings from being carried into the final
    workbook.
    """
    headers = [
        "severity", "entity", "record_id", "issue", "details",
        "recommended_action", "resolution", "reviewer_notes",
    ]
    findings: list[dict[str, object]] = []

    def add(severity: str, entity: str, record_id: str, issue: str,
            details: str, action: str, resolution: str = "Open",
            notes: str = "") -> None:
        findings.append({
            "severity": severity,
            "entity": entity,
            "record_id": record_id,
            "issue": issue,
            "details": details,
            "recommended_action": action,
            "resolution": resolution,
            "reviewer_notes": notes,
        })

    student_numbers = [clean(row.get("student_number")) for row in students]
    for student_number, count in Counter(student_numbers).items():
        if student_number and count > 1:
            add("High", "Student", student_number, "Duplicate student number",
                f"The final roster contains {count} rows with this student number.",
                "Correct the final roster before import.")
    emails = [clean(row.get("email")).lower() for row in students]
    for email, count in Counter(emails).items():
        if email and count > 1:
            add("High", "Student", email, "Duplicate student email",
                f"The final roster produces {count} rows with this email.",
                "Correct the final roster before import.")

    assignments_by_key = {clean(row.get("assignment_key")): row for row in assignments}
    for assignment in assignments:
        key = clean(assignment.get("assignment_key"))
        teacher_id = clean(assignment.get("teacher_user_id"))
        teacher_name = clean(assignment.get("teacher_name"))
        if not teacher_id or not teacher_name or teacher_name.lower() in {
            "new science teacher", "for hire faculty", "tba", "tbd",
        }:
            add(
                "High", "Teaching Assignment", key,
                "Unresolved teacher identity",
                f"{clean(assignment.get('subject_long_name'))} / {clean(assignment.get('canvas_section_name'))}",
                "Confirm an actual teacher or exclude the assignment under the substitution policy.",
            )

    by_pair: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for assignment in assignments:
        by_pair[(clean(assignment.get("section_code")), clean(assignment.get("subject_code")))].append(assignment)
    for (section_code, subject_code), rows in by_pair.items():
        if len(rows) > 1:
            add(
                "Low", "Teaching Assignment",
                f"{section_code}|{subject_code}",
                "Shared class excluded by policy",
                f"{len(rows)} teacher rows exist for this section-subject pair.",
                "Do not create a student evaluation assignment until the class has one eligible teacher.",
                "Resolved",
                "Shared classes are excluded from evaluation in this release.",
            )

    for assignment in student_assignments:
        key = clean(assignment.get("assignment_key"))
        if key not in assignments_by_key:
            add(
                "High", "Student Assignment",
                f"{clean(assignment.get('student_number'))}|{key}",
                "Missing teaching assignment",
                f"{clean(assignment.get('subject_long_name'))} / {clean(assignment.get('canvas_section_name'))}",
                "Remove the student assignment or add a confirmed teaching assignment.",
            )
        if clean(assignment.get("section_code")) not in section_codes:
            add(
                "High", "Student Assignment",
                f"{clean(assignment.get('student_number'))}|{key}",
                "Unknown section code",
                clean(assignment.get("section_code")),
                "Correct the section mapping before import.",
            )
    return headers, findings


def build(source: Path, enrollment: Path, output: Path) -> dict[str, int]:
    if not source.is_file():
        raise SystemExit(f"Source reconciliation workbook does not exist: {source}")
    if not enrollment.is_file():
        raise SystemExit(f"Final enrollment workbook does not exist: {enrollment}")

    workbook = load_workbook(source)
    enrollment_book = load_workbook(enrollment, read_only=True, data_only=True)
    try:
        section_headers, section_rows = records(workbook["Sections"])
        old_student_headers, old_student_rows = records(workbook["Students"])
        assignment_headers, assignment_rows = records(workbook["Teaching Assignments"])
        roster_headers, roster_rows = records(enrollment_book["As of August 25"])
    finally:
        enrollment_book.close()

    section_by_source: dict[str, dict[str, object]] = {}
    for row in old_student_rows:
        source_name = clean(row.get("source_k12_section"))
        if source_name:
            candidate = {
                "canvas_section_name": clean(row.get("canvas_section_name")),
                "section_code": canonical_section_code(
                    row.get("canvas_section_name"), row.get("section_code")
                ),
            }
            previous = section_by_source.setdefault(source_name, candidate)
            if previous != candidate:
                raise SystemExit(f"Ambiguous existing section mapping: {source_name}")

    # The final workbook uses the same official labels as the SIS exports, but
    # the JHS/GAS rows are explicitly guarded so they cannot be lost again.
    explicit = {
        "GRADE 7 Sec-1 (S.Y.26-27)": ("Grade 7-1", "G07-1"),
        "GRADE 7 Sec-2 (S.Y.26-27)": ("Grade 7-2", "G07-2"),
        "GRADE 8 Sec-1 (S.Y.26-27)": ("Grade 8-1", "G08-1"),
        "GRADE 8 Sec-2 (S.Y.26-27)": ("Grade 8-2", "G08-2"),
        "GRADE 9 Sec-1 (S.Y.26-27)": ("Grade 9-1", "G09-1"),
        "GRADE 9 Sec-2 (S.Y.26-27)": ("Grade 9-2", "G09-2"),
        "GRADE 10 Sec-1 (S.Y.26-27)": ("Grade 10-1", "G10-1"),
        "GRADE 10 Sec-2 (S.Y.26-27)": ("Grade 10-2", "G10-2"),
        "12GAS-1A (S.Y. 26-27)": ("12GAS-1A", "12G01a"),
        "12GAS-1B (S.Y. 26-27)": ("12GAS-1B", "12G01b"),
    }
    section_by_source.update(
        {source_name: {"canvas_section_name": canvas, "section_code": code} for source_name, (canvas, code) in explicit.items()}
    )

    # Normalize all source sheets before generating dependent rows. This also
    # repairs duplicate legacy codes in Sections and Teaching Assignments.
    for row in section_rows:
        row["section_code"] = canonical_section_code(
            row.get("canvas_section_name"), row.get("section_code")
        )
    for row in assignment_rows:
        old_key = clean(row.get("assignment_key"))
        row["section_code"] = canonical_section_code(
            row.get("canvas_section_name"), row.get("section_code")
        )
        key_parts = old_key.split("|", 2)
        if len(key_parts) == 3:
            key_parts[1] = re.sub(
                r"_(?:11GE|11GS|11ES)\d{2}$",
                "_" + row["section_code"],
                key_parts[1],
            )
            row["assignment_key"] = "|".join(key_parts)
    for mapping in section_by_source.values():
        mapping["section_code"] = canonical_section_code(
            mapping.get("canvas_section_name"), mapping.get("section_code")
        )

    replace_sheet(workbook, "Sections", section_headers, section_rows)

    # The Canvas course for World Religions carries multiple teachers at the
    # course level. The administrative schedule provides the authoritative
    # section-level allocation, so retain only the scheduled teacher per class.
    world_religions_teacher_by_section = {
        "12H01a": "H202290091",  # Bandolin Jr., Arthur
        "12H01b": "H201902978",  # Basa, Audrey Jeremae
        "12H02a": "H201902978",  # Basa, Audrey Jeremae
        "12H02b": "H201902978",  # Basa, Audrey Jeremae
        "12H03b": "H202290091",  # Bandolin Jr., Arthur
    }
    filtered_assignments: list[dict[str, object]] = []
    for row in assignment_rows:
        section_code = clean(row.get("section_code"))
        if clean(row.get("subject_code")) == "WR" and section_code in world_religions_teacher_by_section:
            if clean(row.get("teacher_user_id")) != world_religions_teacher_by_section[section_code]:
                continue
            row["review_notes"] = "Section teacher verified against 03 Teachers Schedule, SOCSCI (V.2)."
        filtered_assignments.append(row)
    assignment_rows = filtered_assignments

    # The schedule lists two WR sections that were absent from the Canvas
    # assignment snapshot even though the final roster contains students.
    # Materialize those schedule-confirmed rows from the existing WR course
    # template so their students are not silently left without WR evaluation.
    wr_template = next(
        (row for row in assignment_rows if clean(row.get("subject_code")) == "WR"),
        None,
    )
    teacher_lookup = {
        clean(row.get("teacher_user_id")): row
        for row in records(workbook["Teachers"])[1]
    }
    wr_schedule_rows = {
        "12H02a": "H201902978",  # Basa, Audrey Jeremae
        "12H03b": "H202290091",  # Bandolin Jr., Arthur
    }
    if wr_template is not None:
        for section_code, teacher_id in wr_schedule_rows.items():
            if any(
                clean(row.get("section_code")) == section_code
                and clean(row.get("subject_code")) == "WR"
                for row in assignment_rows
            ):
                continue
            row = dict(wr_template)
            teacher = teacher_lookup[teacher_id]
            section_name = next(
                clean(section.get("canvas_section_name"))
                for section in section_rows
                if clean(section.get("section_code")) == section_code
            )
            course_id = clean(row.get("canvas_course_id"))
            row.update({
                "assignment_key": f"SHS|{course_id}_{section_code}|{course_id}|{teacher_id}",
                "canvas_section_name": section_name,
                "section_code": section_code,
                "teacher_user_id": teacher_id,
                "teacher_code": clean(teacher.get("teacher_code")),
                "teacher_name": clean(teacher.get("teacher_name")),
                "teacher_email": clean(teacher.get("teacher_email")),
                "canvas_section_id": f"{course_id}_{section_code}",
                "review_notes": "Section teacher verified against 03 Teachers Schedule, SOCSCI (V.2).",
            })
            assignment_rows.append(row)
    replace_sheet(workbook, "Teaching Assignments", assignment_headers, assignment_rows)

    missing = sorted({clean(row.get("K12 - Section")) for row in roster_rows} - set(section_by_source))
    if missing:
        raise SystemExit("Final roster sections have no reconciliation mapping: " + "; ".join(missing))

    students: list[dict[str, object]] = []
    for row in roster_rows:
        student_number = clean(row.get("Student Number"))
        source_name = clean(row.get("K12 - Section"))
        mapping = section_by_source[source_name]
        grade = grade_number(row.get("Year Level"))
        first = clean(row.get("First Name"))
        middle = clean(row.get("Middle Name"))
        last = clean(row.get("Last Name"))
        students.append({
            "student_number": student_number,
            "canvas_user_id": f"H{student_number}",
            "full_name": " ".join(part for part in (first, middle, last) if part),
            "first_name": first,
            "middle_name": middle,
            "last_name": last,
            "email": f"{student_number}@feuhighschool.edu.ph",
            "school_level": "JHS" if grade <= 10 else "SHS",
            "grade_level": grade,
            "canvas_section_name": mapping["canvas_section_name"],
            "section_code": mapping["section_code"],
            "source_k12_section": source_name,
            "review_status": "Matched",
            "review_notes": "Final Aug 25 enrollment roster",
        })

    by_section: dict[str, list[dict[str, object]]] = defaultdict(list)
    for assignment in assignment_rows:
        by_section[clean(assignment.get("section_code"))].append(assignment)
    assignments_by_key = {clean(row.get("assignment_key")): row for row in assignment_rows}
    subject_counts = Counter(
        (clean(row.get("section_code")), clean(row.get("subject_code"))) for row in assignment_rows
    )
    eligible = {
        clean(row.get("assignment_key"))
        for row in assignment_rows
        if subject_counts[(clean(row.get("section_code")), clean(row.get("subject_code")))] == 1
    }
    student_assignments: list[dict[str, object]] = []
    for student in students:
        for assignment in by_section[clean(student["section_code"])]:
            key = clean(assignment.get("assignment_key"))
            if key in eligible:
                student_assignments.append({
                    "student_number": student["student_number"],
                    "assignment_key": key,
                    "school_level": student["school_level"],
                    "canvas_section_name": student["canvas_section_name"],
                    "section_code": student["section_code"],
                    "subject_code": assignment.get("subject_code"),
                    "subject_long_name": assignment.get("subject_long_name"),
                    "teacher_user_id": assignment.get("teacher_user_id"),
                    "teacher_name": assignment.get("teacher_name"),
                    "review_status": "Ready",
                })

    student_headers = old_student_headers
    student_assignment_headers, _ = records(workbook["Student Assignments"])
    replace_sheet(workbook, "Students", student_headers, students)
    replace_sheet(workbook, "Student Assignments", student_assignment_headers, student_assignments)

    qc_headers, qc_rows = fresh_qc(
        students,
        assignment_rows,
        student_assignments,
        {clean(row.get("section_code")) for row in section_rows},
    )
    replace_sheet(workbook, "QC Issues", qc_headers, qc_rows)

    if "Review Guide" in workbook.sheetnames:
        guide = workbook["Review Guide"]
        metric_counts = {
            "Distinct Canvas section names": len({clean(row.get("canvas_section_name")) for row in section_rows}),
            "Subject codes": len({clean(row.get("subject_code")) for row in records(workbook["Subjects"])[1]}),
            "Teacher Canvas IDs": len({clean(row.get("teacher_user_id")) for row in records(workbook["Teachers"])[1]}),
            "Teaching assignments": len(assignment_rows),
            "Students": len(students),
            "Student-assignment links": len(student_assignments),
            "QC issue rows": len(qc_rows),
            "Students missing email": sum(not clean(row.get("email")) for row in students),
        }
        guide.delete_rows(1, guide.max_row)
        guide.append(["FEU High School Teacher Performance Evaluation Roster Review"])
        guide.append(["Private working file; identifiable student data must remain outside Git."])
        guide.append(["Review metric", "Count"])
        for label, count in metric_counts.items():
            guide.append([label, count])
        guide.append(["Policy and source notes"])
        guide.append(["Student sections", "The Aug 25 K12 - Section value is authoritative for each student."])
        guide.append(["Section codes", "Application section_code values are canonical; SIS and Canvas identifiers are retained for traceability."])
        guide.append(["Teacher identity", "Teacher Canvas IDs are identity keys. Teacher codes are source aliases and are not unique identity keys."])
        guide.append(["Shared classes", "Shared teaching rows are retained for review, but student assignments exclude shared classes."])
        guide.append(["World Religions", "Section-level teachers follow 03 Teachers Schedule, SOCSCI (V.2); Canvas course-level co-enrollments were not used to assign both teachers."])
        guide.append(["Substitutions", "Daily substitutes are excluded; only documented long-term substitutes may be evaluated."])
        guide.append(["Import status", "Blocked until all High QC findings are resolved."])

    if "Source Manifest" in workbook.sheetnames:
        manifest = workbook["Source Manifest"]
        manifest.append([
            enrollment.name, "JHS/SHS", "Authoritative final student roster",
            len(roster_rows), "2026-08-25",
            "Source retained outside Git; K12 - Section controls student placement",
        ])
        manifest.append([
            "sis_export_csv_25_Aug_2026_239520260825-528212-88l3po.csv", "SHS",
            "Latest Canvas SIS enrollment snapshot used for reconciliation",
            "Not loaded by this builder", "2026-08-25",
            "Source retained outside Git; does not override the final student roster",
        ])

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return {
        "students": len(students),
        "student_assignments": len(student_assignments),
        "eligible_assignments": len(eligible),
        "excluded_shared_assignments": len(assignment_rows) - len(eligible),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--enrollment", type=Path, default=DEFAULT_ENROLLMENT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    counts = build(args.source, args.enrollment, args.output)
    print(f"Created {args.output}")
    print("; ".join(f"{key}={value}" for key, value in counts.items()))


if __name__ == "__main__":
    main()
