"""Provision a workbook-based beta cohort without changing the alpha fixture.

The workbook remains local and private. This command reads its normalized
Teaching Assignments sheet, selects approved assignments for named beta
teachers and sections, and creates one synthetic student per selected section.
"""

from __future__ import annotations

import argparse
import csv
import getpass
import hashlib
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from supabase import Client, create_client
from supabase.lib.client_options import SyncClientOptions
from supabase_auth.errors import AuthApiError

try:
    from scripts.provision_alpha import one, secure_password, upsert_one
except ModuleNotFoundError:  # Direct execution from the scripts directory.
    from provision_alpha import one, secure_password, upsert_one

DEFAULT_WORKBOOK = (
    Path("outputs")
    / "ucsp_roster_reconciliation_20260824"
    / "FEU_HS_Teacher_Performance_Evaluation_Roster_UCSP_Reconciled.xlsx"
)
DEFAULT_SECTIONS = ("G07-1", "11HS08", "12S03b", "11PSY02", "12S04a", "11AF03", "11HS06")
DEFAULT_STUDENT_START = 2050000000
BETA_PERIOD_CODE = "BETA-2026-01"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--url", default=os.getenv("SUPABASE_URL", ""))
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--teacher", action="append", required=True, help="Exact workbook teacher name; repeat for each teacher.")
    parser.add_argument("--section", action="append", default=list(DEFAULT_SECTIONS), help="Workbook section code; repeat to override the default seven sections.")
    parser.add_argument("--open-days", type=int, default=7, choices=range(1, 31), metavar="1-30")
    return parser.parse_args()


def workbook_assignments(path: Path, teacher_names: set[str], section_codes: set[str]) -> tuple[list[dict[str, Any]], set[str]]:
    if not path.is_file():
        raise SystemExit(f"Workbook does not exist: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Teaching Assignments"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values, ())]
        rows = [dict(zip(headers, row)) for row in values if any(value not in (None, "") for value in row)]
    finally:
        workbook.close()

    selected = [
        row for row in rows
        if str(row.get("teacher_name") or "") in teacher_names
        and str(row.get("section_code") or "") in section_codes
        and str(row.get("review_status") or "").casefold() in {"ready", "matched"}
    ]
    found_teachers = {str(row.get("teacher_name") or "") for row in selected}
    missing = sorted(teacher_names - found_teachers)
    if missing:
        raise SystemExit("No Ready assignment found for: " + ", ".join(missing))
    if not selected:
        raise SystemExit("No Ready workbook assignments matched the selected teachers and sections.")
    return selected, found_teachers


def main() -> None:
    args = parse_args()
    url = args.url.strip()
    if not url.startswith("https://") or ".supabase.co" not in url:
        raise SystemExit("Provide the hosted Supabase project URL with --url or SUPABASE_URL.")
    teacher_names = set(args.teacher)
    section_codes = set(args.section)
    assignments, _ = workbook_assignments(args.workbook, teacher_names, section_codes)
    selected_sections = sorted({str(row["section_code"]) for row in assignments})
    if len(selected_sections) > 10:
        raise SystemExit("The selected beta cohort would require more than 10 synthetic students.")

    secret_key = getpass.getpass("Supabase secret key (input hidden): ").strip()
    if not secret_key.startswith("sb_secret_"):
        raise SystemExit("A current sb_secret_ key is required for beta provisioning.")
    fingerprint = hashlib.sha256(secret_key.encode("utf-8")).hexdigest()[:8]
    print(f"Secret key received ({len(secret_key)} characters; fingerprint {fingerprint}).")
    client = create_client(
        url,
        secret_key,
        options=SyncClientOptions(auto_refresh_token=False, persist_session=False),
    )
    try:
        users = client.auth.admin.list_users(page=1, per_page=1000)
    except AuthApiError as exc:
        raise SystemExit(f"Supabase administrator authentication failed: {exc}") from None
    planned_emails = {f"beta.student.{index:02d}@example.invalid" for index in range(1, len(selected_sections) + 1)}
    existing = {str(user.email).casefold() for user in users if user.email}
    duplicates = sorted(email for email in planned_emails if email.casefold() in existing)
    if duplicates:
        raise SystemExit("Beta accounts already exist; no changes were made: " + ", ".join(duplicates))

    now = datetime.now(timezone.utc)
    period = upsert_one(client, "evaluation_periods", {
        "code": BETA_PERIOD_CODE,
        "academic_year": "2026-2027",
        "term": "Workbook Beta Test",
        "opens_at": (now - timedelta(minutes=5)).isoformat(),
        "closes_at": (now + timedelta(days=args.open_days)).isoformat(),
        "status": "open",
    }, "code")
    for level in ("JHS", "SHS"):
        bank = one(client.table("question_banks").select("id").eq("code", f"faculty-evaluation-{level.lower()}").eq("version", 1).eq("status", "published").execute(), f"Published {level} question bank was not found.")
        client.table("evaluation_period_instruments").upsert({"evaluation_period_id": period["id"], "school_level": level, "question_bank_id": bank["id"]}, on_conflict="evaluation_period_id,school_level").execute()

    sections: dict[str, dict[str, Any]] = {}
    assignments_by_section: dict[str, list[int]] = {section: [] for section in selected_sections}
    for row in assignments:
        section_code = str(row["section_code"])
        sections[section_code] = upsert_one(client, "sections", {
            "code": section_code,
            "school_level": row["school_level"],
            "grade_level": int(row["grade_level"]),
            "strand": None,
            "is_active": True,
        }, "code")
        teacher = upsert_one(client, "teachers", {
            "employee_number": str(row["teacher_user_id"] or "").removeprefix("H"),
            "display_name": row["teacher_name"],
            "email": row["teacher_email"],
            "is_active": True,
        }, "employee_number")
        subject = upsert_one(client, "subjects", {
            "code": row["subject_code"],
            "name": row["subject_long_name"],
            "is_active": True,
        }, "code")
        assignment = upsert_one(client, "teaching_assignments", {
            "evaluation_period_id": period["id"],
            "section_id": sections[section_code]["id"],
            "subject_id": subject["id"],
            "teacher_id": teacher["id"],
            "is_active": True,
        }, "evaluation_period_id,section_id,subject_id,teacher_id")
        assignments_by_section[section_code].append(int(assignment["id"]))

    credentials: list[dict[str, Any]] = []
    for index, section_code in enumerate(selected_sections):
        row = next(row for row in assignments if str(row["section_code"]) == section_code)
        password = secure_password()
        email = f"beta.student.{index + 1:02d}@example.invalid"
        display_name = f"Workbook Beta Student {index + 1:02d}"
        created = client.auth.admin.create_user({"email": email, "password": password, "email_confirm": True, "user_metadata": {"display_name": display_name}})
        if created.user is None:
            raise RuntimeError(f"Supabase did not return the created user for {email}.")
        user_id = str(created.user.id)
        client.table("profiles").update({"display_name": display_name, "role": "student", "is_active": True}).eq("id", user_id).execute()
        student_number = str(DEFAULT_STUDENT_START + index)
        client.table("students").insert({"profile_id": user_id, "student_number": student_number, "section_id": sections[section_code]["id"]}).execute()
        client.table("student_assignments").insert([{"student_id": user_id, "teaching_assignment_id": assignment_id, "is_active": True} for assignment_id in assignments_by_section[section_code]]).execute()
        credentials.append({"email": email, "password": password, "display_name": display_name, "student_number": student_number, "section": section_code, "grade_level": int(row["grade_level"]), "assignment_count": len(assignments_by_section[section_code])})

    output_dir = Path("exports")
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"beta_workbook_credentials_{timestamp}.csv"
    descriptor = os.open(output, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    with os.fdopen(descriptor, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(credentials[0]))
        writer.writeheader()
        writer.writerows(credentials)
    print(f"Created {len(credentials)} synthetic beta students across {len(assignments)} workbook assignments.")
    print(f"Credentials saved locally with owner-only permissions: {output}")


if __name__ == "__main__":
    main()
