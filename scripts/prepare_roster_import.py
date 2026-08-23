"""Validate the private roster workbook and prepare a Supabase staging bundle.

This command performs no network requests and never uploads roster data. A
bundle is written only when validation has no errors. The output directory is
expected to remain under the Git-ignored ``exports/`` directory.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "Sections": {
        "school_level",
        "grade_level",
        "canvas_section_name",
        "section_code",
        "review_status",
    },
    "Subjects": {"subject_code", "subject_long_name", "review_status"},
    "Teachers": {
        "teacher_user_id",
        "teacher_name",
        "teacher_email",
        "review_status",
    },
    "Teaching Assignments": {
        "assignment_key",
        "section_code",
        "subject_code",
        "teacher_user_id",
        "review_status",
    },
    "Students": {
        "student_number",
        "full_name",
        "email",
        "section_code",
        "review_status",
    },
    "Student Assignments": {
        "student_number",
        "assignment_key",
        "review_status",
    },
    "QC Issues": {"severity", "record_id", "issue", "resolution"},
}

OUTPUT_FILES = {
    "sections": "sections_stage.csv",
    "teachers": "teachers_stage.csv",
    "subjects": "subjects_stage.csv",
    "students": "students_stage.csv",
    "teaching_assignments": "teaching_assignments_stage.csv",
    "student_assignments": "student_assignments_stage.csv",
}


@dataclass(frozen=True)
class ValidationIssue:
    severity: str
    code: str
    entity: str
    key: str
    message: str
    source_sheet: str = ""
    source_row: int | None = None


@dataclass
class PreparedRoster:
    sections: list[dict]
    teachers: list[dict]
    subjects: list[dict]
    students: list[dict]
    teaching_assignments: list[dict]
    student_assignments: list[dict]
    issues: list[ValidationIssue]

    @property
    def error_count(self) -> int:
        return sum(issue.severity == "error" for issue in self.issues)


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def employee_number(value: object) -> str:
    candidate = text(value)
    if candidate.startswith("H") and candidate[1:].isdigit():
        return candidate[1:]
    return candidate


def workbook_rows(path: Path) -> tuple[dict[str, list[dict]], list[ValidationIssue]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    issues: list[ValidationIssue] = []
    try:
        for sheet_name, required in REQUIRED_COLUMNS.items():
            if sheet_name not in workbook.sheetnames:
                issues.append(
                    ValidationIssue(
                        "error",
                        "MISSING_SHEET",
                        "workbook",
                        sheet_name,
                        f"Required sheet {sheet_name!r} is missing.",
                    )
                )
                sheets[sheet_name] = []
                continue
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            headers = [text(value) for value in next(iterator, ())]
            missing = sorted(required - set(headers))
            if missing:
                issues.append(
                    ValidationIssue(
                        "error",
                        "MISSING_COLUMNS",
                        "workbook",
                        sheet_name,
                        "Missing required columns: " + ", ".join(missing),
                        sheet_name,
                        1,
                    )
                )
            rows = []
            for row_number, values in enumerate(iterator, start=2):
                row = {
                    header: text(value)
                    for header, value in zip(headers, values)
                    if header
                }
                if any(row.values()):
                    row["_source_row"] = row_number
                    rows.append(row)
            sheets[sheet_name] = rows
    finally:
        workbook.close()
    return sheets, issues


def duplicate_issues(
    rows: Iterable[Mapping[str, str]],
    fields: tuple[str, ...],
    entity: str,
    code: str,
    sheet: str,
) -> list[ValidationIssue]:
    keys = [tuple(row.get(field, "").casefold() for field in fields) for row in rows]
    duplicates = {key for key, count in Counter(keys).items() if count > 1}
    return [
        ValidationIssue(
            "error",
            code,
            entity,
            "|".join(key),
            f"Duplicate {entity} key for {', '.join(fields)}.",
            sheet,
        )
        for key in sorted(duplicates)
    ]


def prepare_rows(sheets: Mapping[str, list[dict]]) -> PreparedRoster:
    issues: list[ValidationIssue] = []

    for sheet_name in (
        "Sections",
        "Subjects",
        "Teachers",
        "Teaching Assignments",
        "Students",
        "Student Assignments",
    ):
        for row in sheets.get(sheet_name, []):
            status = row.get("review_status", "").casefold()
            if status not in {"ready", "matched"}:
                issues.append(
                    ValidationIssue(
                        "error",
                        "UNAPPROVED_REVIEW_STATUS",
                        sheet_name.casefold().replace(" ", "_"),
                        row.get("student_number")
                        or row.get("assignment_key")
                        or row.get("teacher_user_id")
                        or row.get("subject_code")
                        or row.get("section_code", ""),
                        f"Review status {row.get('review_status', '')!r} is not approved for import.",
                        sheet_name,
                        row.get("_source_row"),
                    )
                )

    sections = [
        {
            "section_code": row.get("section_code", ""),
            "canvas_section_name": row.get("canvas_section_name", ""),
            "school_level": row.get("school_level", "").upper(),
            "grade_level": row.get("grade_level", ""),
            "strand": row.get("strand", ""),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Sections", [])
    ]
    teachers = [
        {
            "employee_number": employee_number(row.get("teacher_user_id")),
            "display_name": row.get("teacher_name", ""),
            "email": row.get("teacher_email", "").lower(),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Teachers", [])
    ]
    subjects = [
        {
            "subject_code": row.get("subject_code", ""),
            "subject_name": row.get("subject_long_name", ""),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Subjects", [])
    ]
    students = [
        {
            "student_number": row.get("student_number", ""),
            "display_name": row.get("full_name", ""),
            "email": row.get("email", "").lower(),
            "section_code": row.get("section_code", ""),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Students", [])
    ]
    teaching_assignments = [
        {
            "assignment_key": row.get("assignment_key", ""),
            "section_code": row.get("section_code", ""),
            "subject_code": row.get("subject_code", ""),
            "teacher_employee_number": employee_number(row.get("teacher_user_id")),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Teaching Assignments", [])
    ]
    student_assignments = [
        {
            "student_number": row.get("student_number", ""),
            "assignment_key": row.get("assignment_key", ""),
            "source_row": row.get("_source_row", ""),
        }
        for row in sheets.get("Student Assignments", [])
    ]

    collections = {
        "section": sections,
        "teacher": teachers,
        "subject": subjects,
        "student": students,
        "teaching_assignment": teaching_assignments,
        "student_assignment": student_assignments,
    }
    for entity, rows in collections.items():
        if not rows:
            issues.append(
                ValidationIssue(
                    "error",
                    "EMPTY_DATASET",
                    entity,
                    "",
                    f"No {entity.replace('_', ' ')} rows were found.",
                )
            )

    issues.extend(duplicate_issues(sections, ("section_code",), "section", "DUPLICATE_SECTION", "Sections"))
    issues.extend(duplicate_issues(teachers, ("employee_number",), "teacher", "DUPLICATE_TEACHER", "Teachers"))
    issues.extend(duplicate_issues(subjects, ("subject_code",), "subject", "DUPLICATE_SUBJECT", "Subjects"))
    issues.extend(duplicate_issues(students, ("student_number",), "student", "DUPLICATE_STUDENT", "Students"))
    issues.extend(duplicate_issues(students, ("email",), "student", "DUPLICATE_STUDENT_EMAIL", "Students"))
    issues.extend(duplicate_issues(teaching_assignments, ("assignment_key",), "teaching_assignment", "DUPLICATE_ASSIGNMENT_KEY", "Teaching Assignments"))
    issues.extend(duplicate_issues(student_assignments, ("student_number", "assignment_key"), "student_assignment", "DUPLICATE_STUDENT_ASSIGNMENT", "Student Assignments"))

    section_codes = {row["section_code"] for row in sections}
    teacher_numbers = {row["employee_number"] for row in teachers}
    subject_codes = {row["subject_code"] for row in subjects}
    assignment_by_key = {row["assignment_key"]: row for row in teaching_assignments}
    student_by_number = {row["student_number"]: row for row in students}

    for row in sections:
        try:
            grade = int(row["grade_level"])
        except (TypeError, ValueError):
            grade = 0
        valid_level = (
            (row["school_level"] == "JHS" and 7 <= grade <= 10)
            or (row["school_level"] == "SHS" and 11 <= grade <= 12)
        )
        if not row["section_code"] or not row["canvas_section_name"] or not valid_level:
            issues.append(
                ValidationIssue(
                    "error", "INVALID_SECTION", "section", row["section_code"],
                    "Section code, name, school level, or grade level is invalid.",
                    "Sections", row["source_row"] or None,
                )
            )

    for row in teachers:
        if not row["employee_number"] or not row["display_name"]:
            issues.append(
                ValidationIssue(
                    "error", "INVALID_TEACHER", "teacher", row["employee_number"],
                    "Teacher employee number and display name are required.",
                    "Teachers", row["source_row"] or None,
                )
            )

    for row in subjects:
        if not row["subject_code"] or not row["subject_name"]:
            issues.append(
                ValidationIssue(
                    "error", "INVALID_SUBJECT", "subject", row["subject_code"],
                    "Subject code and long name are required.",
                    "Subjects", row["source_row"] or None,
                )
            )

    email_pattern = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
    for row in students:
        if not row["student_number"] or not row["display_name"]:
            issues.append(
                ValidationIssue(
                    "error", "INVALID_STUDENT", "student", row["student_number"],
                    "Student number and display name are required.",
                    "Students", row["source_row"] or None,
                )
            )
        if not email_pattern.match(row["email"]):
            issues.append(
                ValidationIssue(
                    "error", "INVALID_STUDENT_EMAIL", "student", row["student_number"],
                    "Student email is missing or invalid.",
                    "Students", row["source_row"] or None,
                )
            )
        if row["section_code"] not in section_codes:
            issues.append(
                ValidationIssue(
                    "error", "UNKNOWN_SECTION", "student", row["student_number"],
                    "Student references a section absent from the Sections sheet.",
                    "Students", row["source_row"] or None,
                )
            )

    assignment_tuples = Counter()
    shared = defaultdict(set)
    for row in teaching_assignments:
        key = row["assignment_key"]
        refs_valid = (
            row["section_code"] in section_codes
            and row["subject_code"] in subject_codes
            and row["teacher_employee_number"] in teacher_numbers
        )
        if not key or not refs_valid:
            issues.append(
                ValidationIssue(
                    "error", "UNKNOWN_ASSIGNMENT_REFERENCE", "teaching_assignment", key,
                    "Teaching assignment has a missing key or unknown section, subject, or teacher.",
                    "Teaching Assignments", row["source_row"] or None,
                )
            )
        assignment_tuples[(row["section_code"], row["subject_code"], row["teacher_employee_number"])] += 1
        shared[(row["section_code"], row["subject_code"])].add(row["teacher_employee_number"])

    for key, count in assignment_tuples.items():
        if count > 1:
            issues.append(
                ValidationIssue(
                    "error", "DUPLICATE_TEACHING_ASSIGNMENT", "teaching_assignment",
                    "|".join(key), "The same section, subject, and teacher occurs more than once.",
                )
            )
    for key, teacher_ids in shared.items():
        if len(teacher_ids) > 1:
            issues.append(
                ValidationIssue(
                    "info", "SHARED_CLASS", "teaching_assignment", "|".join(key),
                    f"Section-subject has {len(teacher_ids)} teachers; student mappings remain teacher-specific.",
                )
            )

    student_subject_teachers = defaultdict(set)
    for row in student_assignments:
        student = student_by_number.get(row["student_number"])
        assignment = assignment_by_key.get(row["assignment_key"])
        if student is None or assignment is None:
            issues.append(
                ValidationIssue(
                    "error", "UNKNOWN_STUDENT_ASSIGNMENT_REFERENCE", "student_assignment",
                    f"{row['student_number']}|{row['assignment_key']}",
                    "Student assignment references an unknown student or teaching assignment.",
                    "Student Assignments", row["source_row"] or None,
                )
            )
            continue
        if student["section_code"] != assignment["section_code"]:
            issues.append(
                ValidationIssue(
                    "error", "CROSS_SECTION_ASSIGNMENT", "student_assignment",
                    f"{row['student_number']}|{row['assignment_key']}",
                    "Student and teaching assignment belong to different sections.",
                    "Student Assignments", row["source_row"] or None,
                )
            )
        student_subject_teachers[(row["student_number"], assignment["subject_code"])].add(
            assignment["teacher_employee_number"]
        )
    for key, teacher_ids in student_subject_teachers.items():
        if len(teacher_ids) > 1:
            issues.append(
                ValidationIssue(
                    "warning", "STUDENT_ASSIGNED_TO_SHARED_TEACHERS", "student_assignment",
                    "|".join(key),
                    f"Student is assigned to {len(teacher_ids)} teachers for one subject; confirm co-teaching is intentional.",
                )
            )

    for row in sheets.get("QC Issues", []):
        resolution = row.get("resolution", "").casefold()
        if resolution == "open" or not resolution:
            issues.append(
                ValidationIssue(
                    "error", "OPEN_WORKBOOK_QC", "qc_issue", row.get("record_id", ""),
                    row.get("issue", "Unresolved workbook QC issue."),
                    "QC Issues", row.get("_source_row"),
                )
            )
        elif resolution == "accepted" and row.get("severity", "").casefold() == "high":
            issues.append(
                ValidationIssue(
                    "error", "ACCEPTED_HIGH_WORKBOOK_QC", "qc_issue", row.get("record_id", ""),
                    "High-severity workbook QC must be resolved, not merely accepted.",
                    "QC Issues", row.get("_source_row"),
                )
            )
        elif resolution == "accepted":
            issues.append(
                ValidationIssue(
                    "warning", "ACCEPTED_WORKBOOK_QC", "qc_issue", row.get("record_id", ""),
                    row.get("issue", "Workbook QC issue was accepted."),
                    "QC Issues", row.get("_source_row"),
                )
            )

    return PreparedRoster(
        sections, teachers, subjects, students,
        teaching_assignments, student_assignments, issues,
    )


def validate_workbook(path: Path) -> PreparedRoster:
    sheets, structural_issues = workbook_rows(path)
    prepared = prepare_rows(sheets)
    prepared.issues[:0] = structural_issues
    return prepared


def write_csv(path: Path, rows: list[dict], batch_code: str) -> None:
    if not rows:
        return
    fields = ["batch_code", *rows[0].keys()]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({"batch_code": batch_code, **row})
    os.chmod(path, 0o600)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Private roster review workbook.")
    parser.add_argument("--batch-code", required=True, help="Unique import batch code.")
    parser.add_argument("--evaluation-period-code", required=True)
    parser.add_argument("--source-date", default=date.today().isoformat())
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("exports/roster_import_bundle"),
        help="Git-ignored destination for the private staging bundle.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.workbook.is_file():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    prepared = validate_workbook(args.workbook)
    counts = {
        key: len(getattr(prepared, key))
        for key in OUTPUT_FILES
    }
    report = {
        "batch_code": args.batch_code,
        "evaluation_period_code": args.evaluation_period_code,
        "source_filename": args.workbook.name,
        "source_sha256": sha256(args.workbook),
        "source_date": args.source_date,
        "counts": counts,
        "errors": prepared.error_count,
        "warnings": sum(issue.severity == "warning" for issue in prepared.issues),
        "info": sum(issue.severity == "info" for issue in prepared.issues),
        "issues": [asdict(issue) for issue in prepared.issues],
    }

    repository_root = Path(__file__).resolve().parents[1]
    output_resolved = args.output_dir.resolve()
    exports_root = (repository_root / "exports").resolve()
    if repository_root == output_resolved or repository_root in output_resolved.parents:
        if exports_root != output_resolved and exports_root not in output_resolved.parents:
            raise SystemExit("Private staging output inside the repository must remain under exports/.")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    report_path = args.output_dir / "validation_report.json"
    report_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    os.chmod(report_path, 0o600)

    if prepared.error_count:
        for filename in (*OUTPUT_FILES.values(), "manifest.json"):
            (args.output_dir / filename).unlink(missing_ok=True)
        raise SystemExit(
            f"Roster validation failed with {prepared.error_count} error(s); "
            f"no staging CSVs were written. Review {report_path}."
        )

    for key, filename in OUTPUT_FILES.items():
        write_csv(args.output_dir / filename, getattr(prepared, key), args.batch_code)
    manifest_path = args.output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(report, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    os.chmod(manifest_path, 0o600)
    print(
        f"Prepared {sum(counts.values()):,} staged rows in {args.output_dir}. "
        f"Warnings: {report['warnings']}; informational findings: {report['info']}."
    )


if __name__ == "__main__":
    main()
