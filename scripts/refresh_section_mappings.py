"""Refresh SIS section codes in a private reconciliation workbook.

The Canvas/enrollment workbook and the SIS section exports use different
labels. This utility resolves the label-to-SIS-code relationship from the
authoritative SIS CSVs, updates the private ``Sections`` sheet, and writes a
new workbook. It never uploads data or changes the source workbook in place.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_RECONCILIATION = Path(
    "outputs/ucsp_roster_reconciliation_20260824/"
    "FEU_HS_Teacher_Performance_Evaluation_Roster_UCSP_Reconciled.xlsx"
)
DEFAULT_JHS = Path(
    "/Users/ronmarccharlesms/Downloads/"
    "sis_export_csv_23_Aug_2026_2388/jhs_sections.csv"
)
DEFAULT_SHS = Path(
    "/Users/ronmarccharlesms/Downloads/"
    "sis_export_csv_23_Aug_2026_2387/shs_sections.csv"
)


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_section_codes(path: Path) -> dict[str, str]:
    """Return exact Canvas/SIS section-name mappings, rejecting ambiguity."""
    if not path.is_file():
        raise SystemExit(f"SIS section export does not exist: {path}")
    grouped: dict[str, set[str]] = defaultdict(set)
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            name = text(row.get("name"))
            section_id = text(row.get("section_id"))
            if not name or not section_id:
                continue
            match = re.search(r"_([^_]+)$", section_id)
            if match:
                grouped[name].add(match.group(1))
    ambiguous = {name: sorted(codes) for name, codes in grouped.items() if len(codes) != 1}
    if ambiguous:
        details = "; ".join(f"{name}: {codes}" for name, codes in sorted(ambiguous.items()))
        raise SystemExit(f"Ambiguous SIS section mapping: {details}")
    return {name: next(iter(codes)) for name, codes in grouped.items()}


def refresh(input_path: Path, output_path: Path, jhs_path: Path, shs_path: Path) -> tuple[int, list[str]]:
    if not input_path.is_file():
        raise SystemExit(f"Reconciliation workbook does not exist: {input_path}")
    mappings = read_section_codes(jhs_path)
    mappings.update(read_section_codes(shs_path))

    workbook = load_workbook(input_path)
    if "Sections" not in workbook.sheetnames:
        raise SystemExit("Reconciliation workbook is missing the Sections sheet.")
    sheet = workbook["Sections"]
    headers = [text(cell.value) for cell in sheet[1]]
    required = {"canvas_section_name", "section_code", "sis_section_code"}
    missing = sorted(required - set(headers))
    if missing:
        raise SystemExit("Sections sheet is missing columns: " + ", ".join(missing))
    positions = {header: index + 1 for index, header in enumerate(headers)}

    changed = 0
    unresolved: list[str] = []
    for row in range(2, sheet.max_row + 1):
        canvas_name = text(sheet.cell(row, positions["canvas_section_name"]).value)
        section_code = text(sheet.cell(row, positions["section_code"]).value)
        if not canvas_name:
            continue
        sis_code = mappings.get(canvas_name)
        if sis_code is None:
            unresolved.append(f"{canvas_name} ({section_code})")
            continue
        cell = sheet.cell(row, positions["sis_section_code"])
        if text(cell.value) != sis_code:
            cell.value = sis_code
            changed += 1

    if unresolved:
        workbook.close()
        raise SystemExit(
            "No SIS mapping was found for: " + "; ".join(sorted(set(unresolved)))
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return changed, sorted(set(mappings))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_RECONCILIATION)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--jhs", type=Path, default=DEFAULT_JHS)
    parser.add_argument("--shs", type=Path, default=DEFAULT_SHS)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output or args.input.with_name(
        args.input.stem + "_SectionsRefreshed.xlsx"
    )
    changed, names = refresh(args.input, output, args.jhs, args.shs)
    print(f"Updated {changed} SIS section codes across {len(names)} mapped labels.")
    print(f"Refreshed workbook: {output}")


if __name__ == "__main__":
    main()
